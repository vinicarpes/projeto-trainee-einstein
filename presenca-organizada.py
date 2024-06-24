import pandas as pd #importando a biblioteca
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

planilhaProd = pd.read_csv('Planilha prod.csv').astype(str) #atribuindo a planilha à variável

planilhaProd.columns =['Situação', 'Identificação do aluno', 'Faltou?', "Data", 'Turno'] #alterando os títulos das colunas

if not planilhaProd.empty: #verificando que não tenha posições vazias
    for index, row in planilhaProd.iterrows(): #laço for que pega o índice e dados de cada linha percorrida 
        if row["Situação"] == "atrasado":
            planilhaProd.at[index, "Situação"] = "Atrasado" 
        else:
            planilhaProd.at[index, "Situação"] = "Sem atraso"

if not planilhaProd.empty: #verificando que não tenha posições vazias
    for index, row in planilhaProd.iterrows():#laço for que pega o índice e dados de cada linha percorrida 
        if row["Faltou?"] == "True":
            planilhaProd.at[index, "Faltou?"] = "Sim" 
        else:
            planilhaProd.at[index, "Faltou?"] = "Não"
  

if not planilhaProd.empty: #verificando que não tenha posições vazias
    for index, row in planilhaProd.iterrows():#laço for que pega o índice e dados de cada linha percorrida 
        if row["Turno"] == "True":
            planilhaProd.at[index, "Turno"] = "17:45 - 20:00" 
        else: planilhaProd.at[index, "Turno"] = "20:15 - 22:00" 

planilhaAlunos = pd.read_csv('ID Alunos.csv').astype(str)

#Criar um dicionário onde as chaves são os IDs dos alunos e os valores são os nomes dos alunos:
id_to_name = pd.Series(planilhaAlunos['nome'].values, index=planilhaAlunos["id"]).to_dict() 

planilhaProd['Nome do Aluno'] = planilhaProd['Identificação do aluno'].map(id_to_name) 

colunas_ordenadas = ['Data', 'Identificação do aluno', 'Nome do Aluno', 'Situação', 'Faltou?', 'Turno'] 
planilhaProd = planilhaProd[colunas_ordenadas] #trocando os valores das linhas de acordo com as posições colunas_ordenadas  

planilhaProd['Data'] = pd.to_datetime(planilhaProd['Data']) #convertendo os valores da coluna data para o formato de data +horário
planilhaProd['Data'] = planilhaProd['Data'].dt.strftime('%d/%m/%Y')#formatando as ordens e quantidade de valores que aparecem

planilhaProd.to_csv('Planilha-presenca-dados-filtrados.csv', sep=';', index=False, encoding='utf-8-sig')

planilhaAlunos.columns = ["ID", 'Nome'] #alterando os valores das colunas da variável
planilhaAlunos['Frequência (%)'] = "" #criando uma coluna vazia
planilhaAlunos['Total de aulas'] = ""
planilhaAlunos['Total de faltas'] = ""

faltas = planilhaProd[planilhaProd['Faltou?']=="Sim"]

# Criar um dicionário onde as chaves são os IDs dos alunos e os valores são as quantidades de faltas
quantidade_faltas = faltas['Identificação do aluno'].value_counts().to_dict() 
quantidade_ids = planilhaProd["Identificação do aluno"].value_counts().to_dict()

planilhaAlunos['Total de aulas'] = planilhaAlunos['ID'].map(quantidade_ids).fillna(0).astype(int)
# Mapear o número de faltas para os alunos em planilhaAlunos
planilhaAlunos['Total de faltas'] = planilhaAlunos["ID"].map(quantidade_faltas).fillna(0).astype(int)

planilhaAlunos['Frequência (%)'] = (((planilhaAlunos['Total de aulas'] -planilhaAlunos['Total de faltas'])/planilhaAlunos['Total de aulas'])* 100).round(2)

#Agrupa as faltas por "Identificação do aluno" e cria uma lista de datas para cada aluno, armazenando o resultado em um dicionário datas_faltas.
datas_faltas = faltas.groupby('Identificação do aluno')['Data'].apply(list).to_dict()
turnos_faltas = faltas.groupby("Identificação do aluno")['Turno'].apply(list).to_dict() 


colunas_ordenadas = [ "ID", 'Nome', 'Total de aulas', 'Total de faltas', 'Frequência (%)']
planilhaAlunos=planilhaAlunos[colunas_ordenadas]

max_faltas = max(map(len, datas_faltas.values())) if datas_faltas else 0
for i in range(max_faltas):
    planilhaAlunos[f'Falta {i + 1}'] = planilhaAlunos["ID"].map(lambda x: datas_faltas.get(x, [])[i] if len(datas_faltas.get(x, [])) > i else '') + " " + planilhaAlunos["ID"].map(lambda x: turnos_faltas.get(x, [])[i] if len(turnos_faltas.get(x, []))> i else "")
#print(planilhaAlunos)

planilhaAlunos.to_csv('Frequência-dos-alunos.csv', sep =';', index=False, encoding='utf-8-sig')

#gerar planilha com: nome, frequecia, mes

dadosBrutos = pd.read_csv('Frequência-dos-alunos.csv', sep=';', encoding='utf-8-sig')
resumoMensal = dadosBrutos[['Nome', 'Frequência (%)']]
resumoMensal.columns=['Nome', 'Frequência absoluta (%)']
resumoMensal['ID'] = planilhaAlunos['ID']

meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
for mes in meses:
    resumoMensal.loc[:, mes] = 0.0

resumoMensal=resumoMensal.copy()

#fazer as frequenncias mensais com base na planilhaProd
planilhaProd['Data'] = pd.to_datetime(planilhaProd['Data'], format='%d/%m/%Y') #convertendo os valores da coluna data para o formato de dia/mes/ano
planilhaProd['Mes'] = planilhaProd['Data'].dt.month_name()  # Create a month column

meses_pt = [
    'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
    'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
]
meses_ing = [
    'January','February','March','April','May','June','July',
    'August','September','October','November','December'
]

meses_dict = dict(zip(meses_ing, meses_pt))

planilhaProd['Mes']=planilhaProd['Mes'].map(meses_dict)


#ocorrencia de cada mes pra cada id
ocorrencia_mes_id = planilhaProd.groupby(['Identificação do aluno', 'Mes']).size().unstack(fill_value=0).to_dict('index')


faltas = planilhaProd[planilhaProd['Faltou?']=="Sim"]

ocorrencia_falta_mes = faltas.groupby(['Mes','Identificação do aluno']).size().unstack(fill_value=0).to_dict()



#agora preciso atrelar o dicionário ao resumoMensal e fazer as operações aritméticas
resumoMensal[meses_pt]=resumoMensal[meses_pt].astype(float)

for index, row in resumoMensal.iterrows():
    aluno_id=row['ID']
    for mes in meses_pt:   
        if aluno_id in ocorrencia_mes_id and mes in ocorrencia_mes_id[aluno_id]:
            frequencia = ((ocorrencia_mes_id[aluno_id][mes] - ocorrencia_falta_mes[aluno_id][mes])/ocorrencia_mes_id[aluno_id][mes])
            resumoMensal.at[index, mes] = round(frequencia*100,2)
        else:
            resumoMensal.loc[:, mes]=0.0

df=pd.DataFrame(resumoMensal)

wb = Workbook()
ws = wb.active
ws.title = 'Resumo Mensal'

colunas_ordenadas = ['ID', 'Nome', 'Frequência absoluta (%)'] + meses_pt
resumoMensal = resumoMensal[colunas_ordenadas]

# Convertendo as colunas de meses para float, se necessário
resumoMensal[meses_pt] = resumoMensal[meses_pt].astype(float)

# Criando um arquivo Excel com openpyxl
wb = Workbook()
ws = wb.active
ws.title = 'Resumo Mensal'

# Copiando os dados do DataFrame para a planilha do Excel
# Usamos dataframe_to_rows diretamente com as colunas ordenadas
rows = dataframe_to_rows(resumoMensal, index=False, header=True)
for r_idx, row in enumerate(rows, 1):
    ws.append(row)

# Aplicando formatação condicional para todas as colunas de meses
for col in range(4, ws.max_column + 1):  # Percorre todas as colunas de 'Janeiro' até 'Dezembro'
    for row in range(2, ws.max_row + 1):  # Percorre todas as linhas com dados (excluindo o cabeçalho)
        if ws.cell(row=row, column=col).value is not None and 0<float(ws.cell(row=row, column=col).value) < 75:
            ws.cell(row=row, column=col).font = Font(color="FF0000")  # Define a cor da fonte para vermelho

# Aplicando formatação condicional para a coluna 'Frequência absoluta (%)'
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=3).value is not None and 0<float(ws.cell(row=row, column=3).value) < 75:
        ws.cell(row=row, column=3).font = Font(color="FF0000")  # Define a cor da fonte para vermelho
print(resumoMensal)
# Salvando o arquivo Excel
wb.save('resumo_mensal_estilizado.xlsx')